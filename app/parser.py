import json
from openpyxl import load_workbook, worksheet
from openpyxl.cell.cell import MergedCell
import gdown
from config import GOOGLE_DRIVE_URL


def download_schedule_file(share_link, output_path='schedule_file.xlsx'):
    # Извлекаем ID файла из ссылки
    file_id = share_link.split('/')[5]
    
    # Формируем прямую ссылку для скачивания
    download_url = f"https://drive.google.com/uc?id={file_id}"
    try:
        # Скачиваем файл
        gdown.download(url=download_url, output=output_path, quiet=False, fuzzy=True)
    except Exception as e:
        print(f"Ошибка при скачивании файла: {e}")
        return

def increase_column_index(index:str, increase_value:int) -> str:
    try:
        letters = ''.join(filter(str.isalpha, index))
        numbers = ''.join(filter(str.isdigit, index))
        numbers = int(numbers) + increase_value
        return letters + str(numbers)
    except Exception as e:
        print(f"Ошибка при изменении индекса:{e}")
        return

def get_cell_value(ws: worksheet, cell_coordinate: str):
    try:
        cell = ws[cell_coordinate]
        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if cell_coordinate in merged_range:
                    top_left_coordinate = merged_range.start_cell.coordinate
                    return ws[top_left_coordinate].value
        else:
            return ws[cell_coordinate].value
    except Exception as e:
        print(f"Error getting value for cell {cell_coordinate}: {e}")
        return None
    
def parse_schedule_column(ws: worksheet, start_index:str):
    index = start_index
    days_dict = {0:'понедельник', 1:'вторник', 2:'среда', 3:'четверг', 4:'пятница', 5:'суббота'}
    week_subjects = []
    for i in range(6):
        day_subjects = parse_schedule_day(ws, index)
        week_subjects.append(
            {
                'day': days_dict.get(i),
                'subjects': day_subjects
            }
        )
        index = increase_column_index(index, 11)
    return week_subjects

def parse_schedule_day(ws:worksheet, start_index:str):
    current_class = 1
    index1 = start_index
    index2 = increase_column_index(start_index, 1)
    day_subjects = []
    for i in range(5):
        cell1_value = get_cell_value(ws, index1)
        cell2_value = get_cell_value(ws, index2)
        if cell1_value == cell2_value:
            subject = '-' if cell1_value is None else cell1_value
            day_subjects.append(
                {
                    'class': current_class,
                    'subject': subject,
                    'numerator': False,
                    'denominator': False,
                    'common': True
                }
            )
        else:
            subject1 = '-' if cell1_value is None else cell1_value
            subject2 = '-' if cell2_value is None else cell2_value
            day_subjects.append(
                {
                    'class': current_class,
                    'subject': subject1,
                    'numerator': True,
                    'denominator': False,
                    'common': False
                }
            )
            day_subjects.append(
                {
                    'class': current_class,
                    'subject': subject2,
                    'numerator': False,
                    'denominator': True,
                    'common': False
                }
            )

        index1 = increase_column_index(index1, 2)
        index2 = increase_column_index(index2, 2)
        current_class += 1
    
    return day_subjects

def parse_schedule_by_groups(input_filename="schedule_file.xlsx", worksheet_name="Бакалавры и магистры"):
    wb = load_workbook(input_filename)
    ws = wb[worksheet_name]
    schedule_by_groups = []
    for row in ws.iter_rows(max_row=1, min_col=5, max_col=36):
        for cell in row:
            index = cell.coordinate
            group_name = ws[index].value
            if group_name is None:
                continue
            index = increase_column_index(index, 1)
            week_subjects = parse_schedule_column(ws, index)
            schedule_by_groups.append(
                {
                    'group_name': group_name,
                    'subjects': week_subjects
                }
            )
    with open('schedule_by_groups.json', 'w', encoding='utf-8') as file:
        json.dump(schedule_by_groups, file, ensure_ascii=False, indent=4)

def parse_classrooms(worksheet_name: str, output_filename: str, input_filename='schedule_file.xlsx'):
    wb = load_workbook(input_filename)
    ws = wb[worksheet_name]
    classrooms = []
    for row in ws.iter_rows(max_row=1, min_col=5, max_col=31):
        for cell in row:
            index = cell.coordinate
            classroom_name = int(ws[index].value)
            index = increase_column_index(index, 1)
            classroom_description = ws[index].value
            index = increase_column_index(index, 1)

            classroom_week = parse_schedule_column(ws, index)
            classrooms.append(
                {
                    'classroom': classroom_name,
                    'description': classroom_description,
                    'subjects': classroom_week
                }
            )
    with open(output_filename, 'w', encoding='utf-8') as file:
        json.dump(classrooms, file, ensure_ascii=False, indent=4)
    
def parse_professors(worksheet_name='Преподаватели', output_filename='professors.json', input_filename='schedule_file.xlsx'):
    wb = load_workbook(input_filename)
    ws = wb[worksheet_name]
    professors_list = []
    for row in ws.iter_rows(max_row=1, min_col=5, max_col=31):
        for cell in row:
            index = cell.coordinate
            professor_name = ws[index].value
            index = increase_column_index(index, 1)
            professor_schedule_week = parse_schedule_column(ws, index)
            professors_list.append(
                {
                    'professor': professor_name,
                    'subjects': professor_schedule_week
                }
            )
    with open(output_filename, 'w', encoding='utf-8') as file:
        json.dump(professors_list, file, ensure_ascii=False, indent=4)

def parse_all_sheets():
    parse_schedule_by_groups()
    parse_classrooms('аудитории', 'classrooms.json')
    parse_classrooms('аудитории (июнь сессия, предзащ', 'classrooms_session.json')
    parse_professors()